import face_recognition
import cv2
from openpyxl import Workbook
from openpyxl.styles import Font,colors
from openpyxl import load_workbook
import datetime
import numpy as np
import pymysql
import pymysql.cursors
import os


def attend():
    now= datetime.datetime.now()
    today=now.day
    month=now.month
    
    try:
        book=load_workbook(str(month)+'.xlsx')
        sheet=book.active
    except Exception as e:
        print("New month has started")
        book=Workbook()
        sheet=book.active
    connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='hci')
    try:
        with connection.cursor() as cursor:
                sql = "SELECT f_name,l_name,present from students  order by f_name,l_name"
                cursor.execute(sql)
                result = cursor.fetchall()
                sql = "UPDATE students SET present=0"
                cursor.execute(sql)
                connection.commit()
                if(len(result)==0):
                    return
                else:
                    known_name=[]
                    for i in range(len(result)):
                        fname=result[i][0]
                        lname=result[i][1]
                        name=fname+" "+lname
                        known_name.append(name)
    finally:
        print("List taken")
    sheet.cell(row=1,column=1).value="Name"
    sheet.cell(row=1,column=1).font=Font(bold=True)
    
    
    
    sheet.cell(row=1,column=int(today)+1).value=str(now.date())
    sheet.cell(row=1,column=int(today)+1).font=Font(bold=True)
    
    
    for i in range(len(known_name)):
        sheet.cell(row=i+2,column=1).value=known_name[i]
    
    for i in range(2,2+len(known_name)):
        sheet.cell(row=i, column=int(today)+1).value = "Absent"
        sheet.cell(row=i, column=int(today)+1).font = Font(color=colors.RED)
    
    
    book.save(str(month)+'.xlsx') 
    video_capture = cv2.VideoCapture(0)
    
    
    path=r"C:\Users\Shaswat\Desktop\HCI -3\static\images\Attendance"
    l=os.listdir(path)
    known_face_encodings =[]
    known_face_names=[]
    for i in range(len(l)):
        s=path+"\\"+l[i]
        img=face_recognition.load_image_file(s)
        known_face_encodings.append(face_recognition.face_encodings(img)[0])
        known_face_names.append(str(i+1))
        
        
        
        
    # Load a sample picture and learn how to recognize it.
#    obama_image = face_recognition.load_image_file(r"C:\Users\Shaswat\Desktop\ALL HCI DOUMENTS\Face-Recognition-Attendance-System-master\1.jpg")
#    obama_face_encoding = face_recognition.face_encodings(obama_image)[0]
#    
#    # Load a second sample picture and learn how to recognize it.
#    biden_image = face_recognition.load_image_file(r"C:\Users\Shaswat\Desktop\ALL HCI DOUMENTS\Face-Recognition-Attendance-System-master\3.jpg")
#    biden_face_encoding = face_recognition.face_encodings(biden_image)[0]
    
    # Create arrays of known face encodings and their names
#    known_face_encodings = [
#        obama_face_encoding,
#        biden_face_encoding
#    ]
#    known_face_names = [
#            "1",
#            "2",
#            "3"
#        ]
    
    
    known_name=["Shaswat Srivastava","Tanya","Shreya"]
    
    
    # Initialize some variables
    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True
    
    while True:
        # Grab a single frame of video
        ret, frame = video_capture.read()
    
        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    
        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = small_frame[:, :, ::-1]
    
        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    
            name = "Unknown"
    
    
            face_names = []
            for face_encoding in face_encodings:
                # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
    
                # # If a match was found in known_face_encodings, just use the first one.
                # if True in matches:
                #     first_match_index = matches.index(True)
                #     name = known_face_names[first_match_index]
    
                # Or instead, use the known face with the smallest distance to the new face
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances)
                if matches[best_match_index]:
                    name = known_face_names[best_match_index]
                    facename=known_name[best_match_index]
                try:
                    with connection.cursor() as cursor:
                        sql = "UPDATE students SET present=1 where f_name=%s and l_name=%s"
                        cursor.execute(sql,(facename.split()[0],facename.split()[1]))
                        connection.commit()
                except:
                    print(facename)
                    print("Name not found")
                    
                    if(name=="Unknown"):
                        pass
                    else:
                        if int(name) in range(1,61):
                            sheet.cell(row=int(name)+1, column=int(today)+1).value = "Present"
                            sheet.cell(row=int(name)+1, column=int(today)+1).font = Font(color=colors.BLUE)
                        else:
                            pass
    
                face_names.append(name)
    
        process_this_frame = not process_this_frame
        print(face_locations)
        if(name=="Unknown"):
            print("."*15+" Not Detecting "+"."*15)
            facename="Unknown"
        else:
            print("."*15+" Detecting "+"."*15)
    
        # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4
    
            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
    
            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, facename, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
    
        # Display the resulting image
        cv2.imshow('Video', frame)
        try:
            book.save(str(now).split(" ")[0]+'.xlsx')
        except:
            print("Permission not given")
    
    
        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            video_capture.release()
            break
    
    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()

