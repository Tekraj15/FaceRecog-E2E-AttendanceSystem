from openpyxl import Workbook
from openpyxl.styles import Font,colors,Color
from openpyxl import load_workbook
import datetime
import pymysql
import pymysql.cursors



def present():
    now= datetime.datetime.now()
    today=now.day
    month=now.month
    
    
    try:
        book=load_workbook(str(month)+'.xlsx')
        sheet=book.active
    except Exception as e:
        print("New month has started")
        book=Workbook()
        sheet=book.active
    connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='hci')
    try:
        with connection.cursor() as cursor:
                sql = "SELECT f_name,l_name,present from students  order by f_name,l_name"
                cursor.execute(sql)
                result = cursor.fetchall()
                if(len(result)==0):
                    return
                else:
                    known_name=[]
                    for i in range(len(result)):
                        fname=result[i][0]
                        lname=result[i][1]
                        name=fname+" "+lname
                        known_name.append(name)
                sql = "UPDATE students SET present=1"
                cursor.execute(sql)
                connection.commit()
    finally:
        connection.close()
    sheet.cell(row=1,column=1).value="Name"
    sheet.cell(row=1,column=1).font=Font(bold=True)
    
    
    
    sheet.cell(row=1,column=int(today)+1).value=str(now.date())
    sheet.cell(row=1,column=int(today)+1).font=Font(bold=True)
    
    
    for i in range(len(known_name)):
        sheet.cell(row=i+2,column=1).value=known_name[i]
    
    for i in range(2,2+len(known_name)):
        sheet.cell(row=i, column=int(today)+1).value = "Present"
        sheet.cell(row=i, column=int(today)+1).font = Font(color=colors.BLUE)
    
    
    book.save(str(now).split(" ")[0]+'.xlsx')

    
